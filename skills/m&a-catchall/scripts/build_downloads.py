#!/usr/bin/env python3
"""Build the xlsx + JSON + CSV downloads for a single-list CatchAll run
(fundraising, M&A, and any single-query event skill).

Shared and data-driven: per-event columns come from the records' own
`enrichment` fields, so one script serves every single-list skill. The agent
never re-writes this logic — it just runs this file.

This script does NO network calls and uses NO API key. The agent pulls the
records through the CatchAll **MCP** (`pull_results`), hands them to this script
via `--input`, and the script only transforms them into files. (In Claude Code
that means MCP-only — never a stored key; on claude.ai the same.)

  --input <file>   A JSON file holding the pulled records. Accepts the raw
                   `pull_results` output ({all_records, candidate_records,
                   valid_records, date_range}), a bare list of records, or an
                   MCP {"result": "<json>"} envelope (auto-unwrapped). If the
                   host saved the MCP response to disk, point --input at it; if
                   not, the agent writes the records there first.

Run metadata via flags: --slug (filenames) --skill --topic --start --end
[--mode base] [--out-dir .].

Writes <slug>.json / .csv / .xlsx and prints the paths to stdout.
"""
import argparse
import csv
import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

DATE_KEYS = ("announcement_date", "event_date", "deal_date", "completion_date", "date")


# --- input -------------------------------------------------------------------

def load_input(path: str) -> dict:
    """Read the records JSON; unwrap an MCP {"result": "<json>"} envelope and
    accept a bare list, {all_records:[...]}, {records:[...]} or {events:[...]}.
    Returns {all_records, candidate_records, valid_records, date_range}."""
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, dict) and isinstance(data.get("result"), str):
        try:
            data = json.loads(data["result"])
        except json.JSONDecodeError:
            pass
    if isinstance(data, list):
        return {"all_records": data, "candidate_records": len(data),
                "valid_records": len(data), "date_range": {}}
    recs = data.get("all_records") or data.get("records") or data.get("events") or []
    return {"all_records": recs,
            "candidate_records": data.get("candidate_records", len(recs)),
            "valid_records": data.get("valid_records", len(recs)),
            "date_range": data.get("date_range") or {}}


# --- links (single source: references/links.json) ----------------------------

def load_links() -> dict:
    """Load the single-source link table shipped next to the skill
    (references/links.json). Fail loudly if absent — the footer URLs live in
    exactly one place and are never reconstructed here."""
    p = Path(__file__).resolve().parent.parent / "references" / "links.json"
    return json.loads(p.read_text(encoding="utf-8"))


def footer_links(links: dict, *, watchlist: bool):
    """(label, url) pairs for the More-with-CatchAll footer, in order. The
    Company Watchlists link is included only on watchlist runs."""
    for link in links["footer_links"]:
        if link.get("watchlist_only") and not watchlist:
            continue
        yield link["label"], link["url"]


# --- record extraction (defensive across shapes) -----------------------------

def domain_of(url: str) -> str:
    if not url:
        return ""
    host = urlparse(url).netloc.lower()
    if not host:  # bare "techcrunch.com/x"
        host = url.split("/", 1)[0].lower()
    return host[4:] if host.startswith("www.") else host


def extract(record: dict) -> dict:
    """Normalize one raw CatchAll record to a clean event dict. Enrichments are
    carried through verbatim, so columns are data-driven per skill."""
    enr = record.get("enrichment") or record.get("enrichments") or {}
    if not isinstance(enr, dict):
        enr = {}
    title = (record.get("record_title") or record.get("title")
             or enr.get("company_name") or enr.get("target")
             or enr.get("target_company") or "Untitled")
    summary = (record.get("event_summary") or record.get("summary")
               or enr.get("event_summary") or enr.get("product_description") or "")
    date = next((enr[k] for k in DATE_KEYS if enr.get(k)), None)
    cits = []
    for c in (record.get("citations") or []):
        link = c.get("link") or c.get("url") or ""
        cits.append({"source": domain_of(link), "url": link,
                     "published_date": c.get("published_date") or c.get("date")})
    if date is None and cits:
        date = cits[0].get("published_date")
    return {"title": title, "date": date, "summary": summary,
            "citations": cits, "enrichments": dict(enr)}


def cell_value(v):
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


# --- builders ----------------------------------------------------------------

def build_json(meta: dict, events: list, path: Path) -> None:
    path.write_text(json.dumps({"meta": meta, "events": events}, indent=2,
                               ensure_ascii=False), encoding="utf-8")


def build_csv(events: list, enr_cols: list, path: Path) -> None:
    head = ["title", "date", "summary", "citation_count", "sources", "urls"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(head + enr_cols)
        for e in events:
            cits = e["citations"]
            w.writerow([
                e["title"], e["date"] or "", e["summary"], len(cits),
                "; ".join(sorted({c["source"] for c in cits if c["source"]})),
                "; ".join(c["url"] for c in cits if c["url"]),
            ] + [str(cell_value(e["enrichments"].get(k))) for k in enr_cols])


def build_xlsx(meta: dict, events: list, enr_cols: list, path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    NAVY = PatternFill("solid", fgColor="1F3A5F")
    TINT = PatternFill("solid", fgColor="F2F4F7")
    HEAD = Font(bold=True, color="FFFFFF")
    LINK = Font(color="2563EB", underline="single")
    TOP = Alignment(vertical="top", wrap_text=False)
    recall = meta.get("recall", {})

    def header(ws, headers):
        for j, h in enumerate(headers, 1):
            c = ws.cell(1, j, h)
            c.fill, c.font = NAVY, HEAD
        ws.freeze_panes = "A2"

    def finish(ws, ncols, nrows):
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(nrows, 1)}"
        for j in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(j)].width = 22

    wb = openpyxl.Workbook()

    ov = wb.active
    ov.title = "Overview"
    ov["A1"] = f"CatchAll — {meta.get('topic', '')}".strip(" —")
    ov["A1"].font = Font(bold=True, size=14)
    ov["A3"] = "One row per event in 'Events'; full citations in 'Citations'."
    ov["A5"] = "More with CatchAll:"
    ov["A5"].font = Font(bold=True)
    row = 6
    links = load_links()
    for label, url in footer_links(links, watchlist=False):
        c = ov.cell(row, 1, label)
        c.hyperlink, c.font = url, LINK
        row += 1
    ov.cell(row, 1, f"Questions? {links['support_email']}")

    ri = wb.create_sheet("Run info")
    win = meta.get("window", {})
    for i, (k, v) in enumerate((
        ("Topic", meta.get("topic", "")),
        ("Window", f"{win.get('start', '')} – {win.get('end', '')}".strip(" –")),
        ("Prepared", meta.get("prepared_at", "")),
        ("Mode", meta.get("mode", "base")),
        ("Web pages scanned", recall.get("total_candidates_scanned", "")),
        ("Events found", recall.get("total_validated_events", "")),
        ("Unique sources", recall.get("unique_sources", "")),
        ("Job ID", meta.get("job_id") or ""),
    ), 1):
        ri.cell(i, 1, k).font = Font(bold=True)
        ri.cell(i, 2, cell_value(v))

    ev = wb.create_sheet("Events")
    head = ["title", "date", "citation_count", "citations", "summary"] + enr_cols
    header(ev, head)
    for i, e in enumerate(sorted(events, key=lambda x: len(x["citations"]), reverse=True), 2):
        cits = e["citations"]
        vals = [e["title"], e["date"] or "", len(cits),
                ", ".join(c["url"] for c in cits if c["url"]), e["summary"]]
        vals += [cell_value(e["enrichments"].get(k)) for k in enr_cols]
        for j, v in enumerate(vals, 1):
            c = ev.cell(i, j, v)
            c.alignment = TOP
            if i % 2 == 0:
                c.fill = TINT
    finish(ev, len(head), len(events) + 1)

    ct = wb.create_sheet("Citations")
    header(ct, ["event_title", "source", "url", "published_date"])
    rows = sorted(((e["title"], c["source"], c["url"], c.get("published_date") or "")
                   for e in events for c in e["citations"]),
                  key=lambda x: (str(x[0]), str(x[1])))
    for i, (t, s, u, pd) in enumerate(rows, 2):
        ct.cell(i, 1, t).alignment = TOP
        ct.cell(i, 2, s).alignment = TOP
        cu = ct.cell(i, 3, u)
        if u:
            cu.hyperlink, cu.font = u, LINK
        ct.cell(i, 4, pd).alignment = TOP
        if i % 2 == 0:
            for col in range(1, 5):
                ct.cell(i, col).fill = TINT
    finish(ct, 4, len(rows) + 1)

    wb.save(path)


# --- main --------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Build CatchAll download files from pulled records.")
    ap.add_argument("--input", required=True, help="JSON file of pulled records (from the MCP).")
    ap.add_argument("--slug", required=True)
    ap.add_argument("--skill", default="")
    ap.add_argument("--topic", default="")
    ap.add_argument("--start", default="")
    ap.add_argument("--end", default="")
    ap.add_argument("--mode", default="base")
    ap.add_argument("--out-dir", dest="out_dir", default=".")
    a = ap.parse_args()

    pull = load_input(a.input)
    events = [extract(r) for r in pull["all_records"]]
    enr_cols = sorted({k for e in events for k in e["enrichments"]})
    unique_sources = len({c["source"] for e in events for c in e["citations"] if c["source"]})
    dr = pull.get("date_range") or {}
    meta = {
        "topic": a.topic,
        "window": {"start": a.start or dr.get("start", ""), "end": a.end or dr.get("end", "")},
        "prepared_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "skill": a.skill,
        "mode": a.mode,
        "job_id": None,
        "recall": {
            "total_candidates_scanned": pull.get("candidate_records", 0),
            "total_validated_events": pull.get("valid_records") or len(events),
            "unique_sources": unique_sources,
        },
    }

    out = Path(a.out_dir)
    out.mkdir(parents=True, exist_ok=True)
    jpath, cpath, xpath = out / f"{a.slug}.json", out / f"{a.slug}.csv", out / f"{a.slug}.xlsx"
    build_json(meta, events, jpath)
    build_csv(events, enr_cols, cpath)

    made = [jpath, cpath]
    try:
        build_xlsx(meta, events, enr_cols, xpath)
        made.append(xpath)
    except ImportError:
        try:
            subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                            "--break-system-packages", "-q"], check=True)
            build_xlsx(meta, events, enr_cols, xpath)
            made.append(xpath)
        except Exception as exc:  # noqa: BLE001
            print(f"build_downloads: xlsx skipped (openpyxl unavailable: {exc})",
                  file=sys.stderr)

    print(f"built {len(events)} events, {unique_sources} sources")
    for p in made:
        print(p.resolve())


if __name__ == "__main__":
    main()
