#!/usr/bin/env python3
"""Render the multi-section report downloads (xlsx + JSON + CSV) for a CatchAll
report skill (competitor-snapshot-catchall, portfolio-monitoring-catchall, and
any sectioned, bucketed skill).

The agent produces the structured report JSON — meta + events-by-bucket +
spotlights — because that shaping needs skill judgment. This script then
deterministically RENDERS that JSON into the three deliverables, so the workbook
is never hand-authored and no data or link is reconstructed at runtime. It is a
pure renderer: it never invents, filters, or reorders data beyond the documented
sort.

  --input <file>   Report JSON: {meta, events:{<bucket>:[...]}, spotlights?}.
                   Accepts an MCP {"result": "<json>"} envelope (auto-unwrapped).
  --slug <slug>    Output basename.
  --out-dir .      Where to write.
  --watchlist      Include the Company Watchlists footer link (watchlist runs).

Writes <slug>.json / .csv / .xlsx and prints the paths to stdout.

Spotlights: meta.spotlights defines each spotlight ({key, name, subtitle,
columns}); top-level `spotlights` maps each key to a list of membership rows
({event_id, <extra column values>}). Every event carries a unique `id`; each
membership row's event_id is resolved against the events and the script FAILS
LOUD on a dangling or duplicate id — an invented spotlight row cannot slip
through. Event data on a spotlight sheet comes from the resolved event; only the
extra columns come from the membership row.
"""
import argparse
import csv
import json
import sys
from pathlib import Path
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    openpyxl = None

NAVY = "1F3A5F"
TINT = "EEF1F5"
BASE_FIELDS = frozenset({"title", "company", "date", "ed_score", "citation_count",
                         "citations", "relation", "summary"})


# --- links (single source: references/links.json) ----------------------------

def load_links() -> dict:
    """Load the single-source link table shipped next to the skill. Fail loudly
    if absent — footer URLs live in exactly one place, never reconstructed."""
    p = Path(__file__).resolve().parent.parent / "references" / "links.json"
    return json.loads(p.read_text(encoding="utf-8"))


def footer_links(links: dict, *, watchlist: bool):
    for link in links["footer_links"]:
        if link.get("watchlist_only") and not watchlist:
            continue
        yield link["label"], link["url"]


# --- input -------------------------------------------------------------------

def load_report(path: str):
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, dict) and isinstance(data.get("result"), str):
        try:
            data = json.loads(data["result"])
        except json.JSONDecodeError:
            pass
    if not isinstance(data, dict) or not isinstance(data.get("events"), dict):
        sys.exit("build_report: input must be an object with an 'events' object "
                 "keyed by bucket.")
    return data.get("meta") or {}, data["events"], data.get("spotlights") or {}


# --- Phase-3 extraction: build events from truly-raw per-bucket pulls ---------

def _enrichment_types(pull: dict) -> dict:
    """name -> declared type, from the pull's own enrichments[] schema."""
    return {e["name"]: e.get("type") for e in (pull.get("enrichments") or [])}


def _company_value(v):
    """A company-type value is {source_text, confidence, metadata:{name, domain_url}}.
    `metadata.name` is the CANONICAL resolved company name; `source_text` is only the
    matched text span, which is routinely a whole sentence ("Operated by EDF, the plant
    has been generating electricity since 1995"). Take the resolved name, falling back
    to the span when resolution returned none — never invent, never drop."""
    if not isinstance(v, dict):
        return v
    return (v.get("metadata") or {}).get("name") or v.get("source_text")


def _extract_enrichments(raw_enr: dict, types: dict) -> dict:
    """Copy each declared enrichment verbatim, resolving company-type objects to the
    company name. Scalars pass through. enrichment_confidence and anything not in the
    schema is dropped."""
    out = {}
    for name, typ in types.items():
        v = raw_enr.get(name)
        out[name] = _company_value(v) if typ == "company" else v
    return out


def _attribution(record: dict):
    """Watchlist attribution: the highest-ed_score connected entity ->
    (company, ed_score, relation). Deterministic; a tie keeps list order."""
    ents = record.get("connected_entities") or []
    if not ents:
        return "", "", ""
    e = ents[max(range(len(ents)), key=lambda i: (ents[i].get("ed_score") or 0))]
    score = e.get("ed_score", "")
    if isinstance(score, float) and score.is_integer():  # CatchAll sends 10.0; show 10 everywhere
        score = int(score)
    return e.get("name", ""), score, e.get("relation", "")


def _event_from_record(record: dict, types: dict) -> dict:
    """One truly-raw CatchAll record -> one event, every field copied verbatim.
    The model never touches these values."""
    company, ed_score, relation = _attribution(record)
    cits = []
    for c in (record.get("citations") or []):
        url = c.get("link") or c.get("url") or ""
        cits.append({"source": domain_of(url), "url": url,
                     "published_date": (c.get("published_date") or "")})
    date = cits[0]["published_date"][:10] if cits and cits[0]["published_date"] else None
    enr = _extract_enrichments(record.get("enrichment") or {}, types)
    # raw has no native summary field: watchlist records use the entity relation
    # (verbatim); otherwise a skill-defined `summary` enrichment fills it. Either
    # way the summary never doubles as an enrichment column.
    summary = relation or (enr.pop("summary", None) or "")
    ev = {
        "id": str(record.get("record_id")),
        "title": record.get("record_title", ""),
        "date": date,
        "summary": summary,
        "citations": cits,
        "enrichments": enr,
    }
    if company or relation or ed_score != "":
        # watchlist attribution — non-watchlist records carry no entity fields
        ev["company"], ev["ed_score"], ev["relation"] = company, ed_score, relation
    return ev


def extract_from_raw(raw_dir: str):
    """Build (meta, events, spotlights) from a directory of verbatim per-bucket
    pull_results dumps + a manifest.json. The model produces only the raw dumps
    (a copy of the MCP output), the spotlight selection (by record_id), and run
    meta — never any event field.

    manifest.json = { meta{entity,window,skill,mode}, bucket_order[],
                      spotlights_meta[], spotlights{key:[{record_id, <extra cols>}]} }
    <bucket>.json  = the verbatim pull_results response for that bucket."""
    d = Path(raw_dir)
    manifest = json.loads((d / "manifest.json").read_text(encoding="utf-8"))
    meta = dict(manifest.get("meta") or {})
    events, per_bucket = {}, {}
    for bucket in manifest["bucket_order"]:
        pull = json.loads((d / f"{bucket}.json").read_text(encoding="utf-8"))
        if isinstance(pull, dict) and isinstance(pull.get("result"), str):
            pull = json.loads(pull["result"])  # unwrap the MCP tool-result envelope
        if not isinstance(pull, dict) or "all_records" not in pull:
            sys.exit(f"build_report: raw/{bucket}.json has no 'all_records' — expected the "
                     f"verbatim pull (or its {{'result': ...}} tool envelope).")
        types = _enrichment_types(pull)
        events[bucket] = [_event_from_record(r, types) for r in (pull.get("all_records") or [])]
        per_bucket[bucket] = {"job_id": pull.get("job_id"),
                              "candidates_scanned": pull.get("candidate_records"),
                              "validated_events": pull.get("valid_records", len(events[bucket]))}
    meta["per_bucket"] = per_bucket
    meta["spotlights"] = manifest.get("spotlights_meta") or []
    rec = meta.setdefault("recall", {})
    rec["total_candidates_scanned"] = sum((pb.get("candidates_scanned") or 0) for pb in per_bucket.values())
    rec["total_validated_events"] = sum(len(v) for v in events.values())
    rec["unique_sources"] = len({c["source"] for evs in events.values() for e in evs
                                 for c in e["citations"] if c["source"]})
    spotlights = {k: [{"event_id": str(m["record_id"]),
                       **{kk: vv for kk, vv in m.items() if kk != "record_id"}} for m in members]
                  for k, members in (manifest.get("spotlights") or {}).items()}
    return meta, events, spotlights


def index_events(events: dict) -> dict:
    """Map event id -> event, assigning a positional id (<bucket>_<n>) when the
    event has none. Enforce global uniqueness so spotlight refs are unambiguous."""
    idx = {}
    for bucket, evs in events.items():
        for i, e in enumerate(evs, 1):
            eid = e.get("id") or f"{bucket}_{i}"
            e["id"] = eid
            if eid in idx:
                sys.exit(f"build_report: duplicate event id {eid!r}.")
            idx[eid] = e
    return idx


# --- helpers -----------------------------------------------------------------

def domain_of(url: str) -> str:
    if not url:
        return ""
    host = urlparse(url).netloc.lower() or url.split("/", 1)[0].lower()
    return host[4:] if host.startswith("www.") else host


def cites(e: dict) -> list:
    out = []
    for c in (e.get("citations") or []):
        url = c.get("url") or c.get("link") or ""
        out.append({"source": c.get("source") or domain_of(url), "url": url,
                    "published_date": c.get("published_date") or c.get("date") or ""})
    return out


def cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def run_is_watchlist(events: dict) -> bool:
    return any(e.get(k) not in (None, "") for evs in events.values() for e in evs
               for k in ("company", "ed_score", "relation"))


def bucket_enr_cols(evs: list) -> list:
    """Enrichment columns for a bucket, sorted, with is_developing forced last."""
    cols = sorted({k for e in evs for k in (e.get("enrichments") or {})
                   if k != "is_developing"})
    if any("is_developing" in (e.get("enrichments") or {}) for e in evs):
        cols.append("is_developing")
    return cols


def data_columns(*, has_wl: bool, extra: list, enr: list) -> list:
    """Standard left-to-right column order (OUTPUT-REPORT.md § column conventions):
    title, [company], date, [spotlight extras], [ed_score], citation_count,
    citations, [relation], summary, [bucket enrichments]."""
    cols = ["title"] + (["company"] if has_wl else []) + ["date"] + list(extra)
    cols += (["ed_score"] if has_wl else []) + ["citation_count", "citations"]
    cols += (["relation"] if has_wl else []) + ["summary"] + list(enr)
    return cols


def row_values(e: dict, cols: list, extra: dict) -> list:
    c = cites(e)
    base = {
        "title": e.get("title", ""),
        # non-watchlist skills carry the company as an enrichment instead
        "company": e.get("company") or (e.get("enrichments") or {}).get("company") or "",
        "date": e.get("date") or "",
        "ed_score": e.get("ed_score", ""),
        "citation_count": len(c),
        "citations": ", ".join(x["url"] for x in c if x["url"]),
        "relation": e.get("relation", ""),
        "summary": e.get("summary", ""),
    }
    enr = e.get("enrichments") or {}
    out = []
    for col in cols:
        if col in base:
            out.append(base[col])
        elif col in extra:
            out.append(cell(extra[col]))
        elif col == "is_developing":
            v = enr.get("is_developing")
            out.append("" if v is None else str(cell(v)).lower())
        elif col in enr:
            out.append(cell(enr[col]))
        else:
            out.append("")
    return out


def by_citations(evs: list) -> list:
    return sorted(evs, key=lambda e: len(cites(e)), reverse=True)


# --- chat digest (display-ready rows the model reads to write the chat) -------

def _source_str(c: list) -> str:
    """First citation domain + count of others — the Sources cell, prebuilt."""
    if not c:
        return ""
    n = len(c) - 1
    return c[0]["source"] if n == 0 else f'{c[0]["source"]} + {n} other{"s" if n > 1 else ""}'


def _event_digest(e: dict, has_wl: bool) -> dict:
    c = cites(e)
    enr = e.get("enrichments") or {}
    d = {"title": e.get("title", ""), "date": e.get("date") or "",
         "sources": _source_str(c), "citation_count": len(c), "enrichments": enr}
    # company: watchlist attribution when there is one, else the company enrichment —
    # same resolution row_values uses, so the digest and the sheets agree
    company = e.get("company") or enr.get("company") or ""
    if company:
        d["company"] = company
    if has_wl:
        d["ed_score"] = e.get("ed_score", "")
    return d


def chat_digest(meta: dict, events: dict, spotlights: dict, idx: dict, has_wl: bool) -> dict:
    """Everything the chat report needs, already shaped: dashboard numbers, each
    bucket's events (sorted, with a prebuilt Sources string + enrichment values),
    and spotlights with their rows RESOLVED. The model reads this instead of
    re-extracting from the JSON — so it writes no ad-hoc parsing code."""
    rec, pb = meta.get("recall") or {}, meta.get("per_bucket") or {}
    buckets = []
    for bucket, evs in events.items():
        info = pb.get(bucket) or {}
        buckets.append({"key": bucket, "candidates_scanned": info.get("candidates_scanned"),
                        "events_found": info.get("validated_events", len(evs)),
                        "events": [_event_digest(e, has_wl) for e in by_citations(evs)]})
    spots = []
    for spec in meta.get("spotlights", []):
        members = spotlights.get(spec.get("key")) or []
        rows = []
        for m in members:
            e = idx.get(m.get("event_id"))
            if e is None:
                continue
            row = _event_digest(e, has_wl)
            row.update({k: v for k, v in m.items() if k != "event_id"})
            rows.append(row)
        if rows:
            # membership order IS the skill's selection order — the rule already
            # sorted it (citations, severity, date, …). Never re-sort here.
            spots.append({"name": spec.get("name", spec.get("key")),
                          "subtitle": spec.get("subtitle", ""), "rows": rows})
    # value counts for any enrichment listed in meta.aggregate_counts, sorted by
    # count — so the chat reads per-value totals instead of counting rows
    aggregates = {}
    for key in (meta.get("aggregate_counts") or []):
        counts = {}
        for evs in events.values():
            for e in evs:
                v = (e.get("enrichments") or {}).get(key)
                if v not in (None, ""):
                    counts[str(v)] = counts.get(str(v), 0) + 1
        if counts:
            aggregates[key] = [{"value": k, "count": n}
                               for k, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))]
    digest = {"dashboard": {"entity": meta.get("entity", ""), "window": meta.get("window", ""),
                            "prepared": meta.get("prepared_at", ""),
                            "mode": (meta.get("mode") or "base").title(),
                            "total_pages_scanned": rec.get("total_candidates_scanned"),
                            "total_events": rec.get("total_validated_events")},
              "buckets": buckets, "spotlights": spots}
    if aggregates:
        digest["aggregates"] = aggregates
    return digest


def safe_title(name: str, used: set) -> str:
    t = "".join(ch for ch in str(name) if ch not in r'[]:*?/\\')[:31] or "Sheet"
    base, n = t, 2
    while t in used:
        suffix = f" {n}"
        t = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(t)
    return t


# --- xlsx --------------------------------------------------------------------

def build_xlsx(meta, events, spotlights, idx, links, watchlist, path):
    head_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="2563EB", underline="single")
    navy = PatternFill("solid", fgColor=NAVY)
    tint = PatternFill("solid", fgColor=TINT)
    top = Alignment(vertical="top", wrap_text=False)
    has_wl = run_is_watchlist(events)
    used = set()

    def sheet(title):
        return wb.create_sheet(safe_title(title, used))

    def write_grid(ws, cols, rows, subtitle=None):
        hrow = 1
        if subtitle:
            ws.cell(1, 1, subtitle).font = Font(italic=True, color="8D8A78")
            hrow = 2
        for j, h in enumerate(cols, 1):
            c = ws.cell(hrow, j, h)
            c.fill, c.font = navy, head_font
        ws.freeze_panes = f"A{hrow + 1}"
        for k, vals in enumerate(rows):
            i = hrow + 1 + k
            for j, v in enumerate(vals, 1):
                c = ws.cell(i, j, cell(v))
                c.alignment = top
                if k % 2:
                    c.fill = tint
        last = get_column_letter(max(len(cols), 1))
        ws.auto_filter.ref = f"A{hrow}:{last}{hrow + len(rows)}"
        for j in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 22

    wb = openpyxl.Workbook()

    # 1. Overview
    ov = wb.active
    ov.title = "Overview"
    ov["A1"] = f"CatchAll — {meta.get('entity', meta.get('topic', ''))}".strip(" —")
    ov["A1"].font = Font(bold=True, size=14)
    ov["A3"] = ("One sheet per section; spotlight sheets highlight cross-cutting "
                "events; full citations in 'Citations'.")
    ov["A5"] = "More with CatchAll:"
    ov["A5"].font = Font(bold=True)
    r = 6
    for label, url in footer_links(links, watchlist=watchlist):
        c = ov.cell(r, 1, label)
        c.hyperlink, c.font = url, link_font
        r += 1
    ov.cell(r, 1, f"Questions? {links['support_email']}")

    # 2. Run info
    ri = sheet("Run info")
    win = meta.get("window", {})
    recall = meta.get("recall", {})
    info = [
        ("Entity", meta.get("entity", meta.get("topic", ""))),
        ("Window", f"{win.get('start', '')} - {win.get('end', '')}".strip(" -")),
        ("Prepared", meta.get("prepared_at", "")),
        ("Mode", meta.get("mode", "base")),
        ("Companies", meta.get("company_count", len(meta.get("companies", [])) or "")),
        ("Web pages scanned", recall.get("total_candidates_scanned", "")),
        ("Events found", recall.get("total_validated_events", "")),
        ("Unique sources", recall.get("unique_sources", "")),
    ]
    for i, (k, v) in enumerate(info, 1):
        ri.cell(i, 1, k).font = Font(bold=True)
        ri.cell(i, 2, cell(v))
    start = len(info) + 2
    hdr = ["Category", "Web pages scanned", "Events found", "Job ID"]
    for j, h in enumerate(hdr, 1):
        c = ri.cell(start, j, h)
        c.fill, c.font = navy, head_font
    per_bucket = meta.get("per_bucket", {})
    for i, bucket in enumerate(events, start + 1):
        pb = per_bucket.get(bucket, {})
        vals = [bucket, pb.get("candidates_scanned", ""),
                pb.get("validated_events", len(events[bucket])), pb.get("job_id", "")]
        for j, v in enumerate(vals, 1):
            ri.cell(i, j, cell(v))
    for j in range(1, 5):
        ri.column_dimensions[get_column_letter(j)].width = 26

    # 3. Spotlight sheets (defs in meta.spotlights, membership in `spotlights`)
    for spec in meta.get("spotlights", []):
        key = spec.get("key")
        members = spotlights.get(key) or []
        if not members:
            continue
        # columns = the full ordered list the skill declares; base fields come
        # from the event, any other column's value comes from the membership row
        cols = spec.get("columns") or data_columns(has_wl=has_wl, extra=[], enr=[])
        resolved = []  # (event, membership-values dict)
        for m in members:
            eid = m.get("event_id")
            if eid not in idx:
                sys.exit(f"build_report: spotlight {key!r} references unknown "
                         f"event_id {eid!r} — refusing to invent a row.")
            resolved.append((idx[eid], {k: v for k, v in m.items() if k != "event_id"}))
        # a declared non-base column that no membership row supplies renders blank —
        # almost always a mistake (forgot the value, or misdeclared a base field)
        for col in cols:
            if col not in BASE_FIELDS and not any(x.get(col) not in (None, "") for _, x in resolved):
                sys.stderr.write(f"build_report: warning — spotlight {key!r} column {col!r} is "
                                 f"empty in every row (missing from the membership rows?).\n")
        # keep the manifest's membership order (the skill's own sort) — see chat_digest
        rows = [row_values(e, cols, extra) for e, extra in resolved]
        write_grid(sheet(spec.get("name", key)), cols, rows, subtitle=spec.get("subtitle"))

    # 4. One sheet per bucket (in the JSON's order; empty buckets get a header)
    for bucket, evs in events.items():
        enr = bucket_enr_cols(evs)
        cols = data_columns(has_wl=has_wl, extra=[], enr=enr)
        rows = [row_values(e, cols, {}) for e in by_citations(evs)]
        write_grid(sheet(bucket), cols, rows)

    # 5. Citations (one row per citation, across the whole run)
    ct = sheet("Citations")
    crows = []
    for bucket, evs in events.items():
        for e in evs:
            for c in cites(e):
                crows.append((bucket, e.get("title", ""), c["source"], c["url"],
                              c["published_date"]))
    crows.sort(key=lambda x: (str(x[0]), str(x[1]), str(x[2])))
    for j, h in enumerate(["category", "event_title", "source", "url", "published_date"], 1):
        c = ct.cell(1, j, h)
        c.fill, c.font = navy, head_font
    ct.freeze_panes = "A2"
    for i, (cat, title, src, url, pd) in enumerate(crows, 2):
        ct.cell(i, 1, cat).alignment = top
        ct.cell(i, 2, title).alignment = top
        ct.cell(i, 3, src).alignment = top
        cu = ct.cell(i, 4, url)
        if url:
            cu.hyperlink, cu.font = url, link_font
        ct.cell(i, 5, pd).alignment = top
        if i % 2:
            for col in range(1, 6):
                ct.cell(i, col).fill = tint
    ct.auto_filter.ref = f"A1:E{max(len(crows) + 1, 1)}"
    for j in range(1, 6):
        ct.column_dimensions[get_column_letter(j)].width = 26

    wb.save(path)


# --- csv ---------------------------------------------------------------------

def build_csv(events, has_wl, path):
    enr_cols = sorted({k for evs in events.values() for e in evs
                       for k in (e.get("enrichments") or {}) if k != "is_developing"})
    has_dev = any("is_developing" in (e.get("enrichments") or {})
                  for evs in events.values() for e in evs)
    head = ["bucket"] + (["company", "ed_score", "relation"] if has_wl else [])
    head += ["title", "date", "citation_count", "citations", "summary"] + enr_cols
    if has_dev:
        head.append("is_developing")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(head)
        for bucket, evs in events.items():
            for e in evs:
                c = cites(e)
                enr = e.get("enrichments") or {}
                row = [bucket]
                if has_wl:
                    row += [e.get("company", ""), e.get("ed_score", ""), e.get("relation", "")]
                row += [e.get("title", ""), e.get("date") or "", len(c),
                        ", ".join(x["url"] for x in c if x["url"]), e.get("summary", "")]
                row += [cell(enr.get(k, "")) for k in enr_cols]
                if has_dev:
                    v = enr.get("is_developing")
                    row.append("" if v is None else str(cell(v)).lower())
                w.writerow(row)


# --- main --------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Render CatchAll report downloads.")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--raw-dir", dest="raw_dir",
                     help="Dir of verbatim per-bucket pull dumps + manifest.json — the script "
                          "extracts every event field from the raw pulls (nothing hand-written).")
    src.add_argument("--input", help="Pre-built report JSON (meta + events-by-bucket + spotlights).")
    ap.add_argument("--slug", required=True)
    ap.add_argument("--out-dir", dest="out_dir", default=".")
    ap.add_argument("--watchlist", action="store_true",
                    help="Include the Company Watchlists footer link (watchlist runs).")
    a = ap.parse_args()

    meta, events, spotlights = extract_from_raw(a.raw_dir) if a.raw_dir else load_report(a.input)
    idx = index_events(events)  # assigns ids + enforces uniqueness
    has_wl = run_is_watchlist(events)

    out = Path(a.out_dir)
    out.mkdir(parents=True, exist_ok=True)
    jpath, cpath, xpath = out / f"{a.slug}.json", out / f"{a.slug}.csv", out / f"{a.slug}.xlsx"

    jpath.write_text(json.dumps({"meta": meta, "events": events, "spotlights": spotlights},
                                indent=2, ensure_ascii=False), encoding="utf-8")
    build_csv(events, has_wl, cpath)
    made = [jpath, cpath]
    if openpyxl is None:
        sys.stderr.write("build_report: openpyxl not installed — wrote JSON + CSV only. "
                         "Run `pip install openpyxl --break-system-packages` for the xlsx.\n")
    else:
        build_xlsx(meta, events, spotlights, idx, load_links(), a.watchlist, xpath)
        made.append(xpath)

    for p in (xpath, jpath, cpath):
        if p in made:
            print(p.resolve())

    # Chat digest — display-ready rows the model reads to write the chat (no
    # re-extraction). File only; echoing it would just double the read.
    digest = chat_digest(meta, events, spotlights, idx, has_wl)
    digest_path = (Path(a.raw_dir) / "digest.json") if a.raw_dir else out / f"{a.slug}.digest.json"
    digest_path.write_text(json.dumps(digest, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nCHAT DIGEST -> {digest_path.resolve()} — read it and write the chat report "
          "from it; do NOT re-open the JSON/CSV to extract rows.")


if __name__ == "__main__":
    main()
