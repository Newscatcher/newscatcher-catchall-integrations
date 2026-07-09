#!/usr/bin/env python3
"""VC pack render: raw CatchAll pulls in, dashboard out.

Reads a JSON document with two raw pull responses + meta, computes all
aggregates (FX-converted, USD-equivalent, capital ratio, mega rounds,
deal stage buckets, sub-sectors, top-3 lists), pre-renders all table
rows and KPI HTML, substitutes into dashboard.html, and emits HTML to
stdout (or --output PATH).

    python render.py input.json [--format html|markdown] [--output PATH]

Input shape (see SKILL.md § Building the two queries):

    {
      "funding_pull": { "candidate_records": int, "valid_records": int,
                        "all_records": [...] },
      "ma_pull":      { "candidate_records": int, "valid_records": int,
                        "all_records": [...] },
      "meta": { "market": str, "location": str, "window_days": int,
                "start_date": str, "end_date": str }
    }

All aggregation happens in Python — Claude's job upstream is to pull both
feeds and call this script. Nothing else.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# Config
# =============================================================================

# Fixed approximate USD->native rates — no live FX call (deterministic; works
# offline, e.g. in claude.ai's sandbox). To convert native->USD: divide by the
# rate. Update periodically; unlisted currencies stay unconverted (shown
# per-deal, left out of the USD totals).
FX_RATES_USD = {
    "USD": 1.0, "EUR": 0.92, "GBP": 0.79, "JPY": 150.0, "INR": 83.0,
    "KRW": 1350.0, "CNY": 7.2, "CAD": 1.36, "AUD": 1.52, "CHF": 0.88,
    "SGD": 1.35, "HKD": 7.8, "BRL": 5.4, "SEK": 10.5, "NOK": 10.7,
    "DKK": 6.9, "AED": 3.67, "ILS": 3.7, "ZAR": 18.5, "MXN": 18.0,
}
FX_RATES_AS_OF = "approx., mid-2026"
MEGA_ROUND_THRESHOLD = 100_000_000


# Records with deal_value_value above this and display "Undisclosed" are
# treated as bogus extractions (target AUM, not deal price). See §5.3.
DEAL_VALUE_BOGUS_THRESHOLD = 50_000_000_000

DEAL_STAGE_META = {
    "early": {"label": "Early-stage", "sublabel": "(Seed, A)",      "color": "#4A8B7E"},
    "mid":   {"label": "Mid-stage",   "sublabel": "(B, C)",         "color": "#2D6B5E"},
    "late":  {"label": "Late-stage",  "sublabel": "(Growth, D+)",   "color": "#173E36"},
    "other": {"label": "Other",       "sublabel": "(unknown)",      "color": "#B5B2AA"},
}

ACQUIRER_TYPE_META = {
    "big_tech":  {"label": "Big Tech",  "color": "#378ADD", "chipClass": "buyer-chip-leader"},
    "strategic": {"label": "Strategic", "color": "#7F77DD", "chipClass": "buyer-chip-same"},
    "financial": {"label": "Financial", "color": "#C49152", "chipClass": "buyer-chip-financial"},
    "other":     {"label": "Other",     "color": "#B5B2AA", "chipClass": "buyer-chip-unclear"},
}

STAGE_BUCKETS = {
    "pre-seed":  "early",
    "preseed":   "early",
    "pre_seed":  "early",
    "seed":      "early",
    "series_a":  "early",
    "series-a":  "early",
    "series_b":  "mid",
    "series-b":  "mid",
    "series_c":  "mid",
    "series-c":  "mid",
    "series_d":  "late",
    "series-d":  "late",
    "series_e":  "late",
    "series-e":  "late",
    "growth":    "late",
}


# =============================================================================
# Helpers — formatting, escaping, etc.
# =============================================================================

def esc(value) -> str:
    """HTML-escape a string for safe interpolation in element text or attrs."""
    if value is None:
        return ""
    return html.escape(str(value), quote=True)


def _strip_zeros(s: str) -> str:
    """Remove trailing zeros and trailing decimal point from a decimal string."""
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def fmt_usd(amount, undisclosed: str = "Undisclosed") -> str:
    """Compact USD: $1.02B, $3.2B, $192M, $50M, $3.5M, $500K, $0, Undisclosed."""
    if amount is None:
        return undisclosed
    a = float(amount)
    if a == 0:
        return "$0"
    if a >= 1_000_000_000:
        s = _strip_zeros(f"{a / 1_000_000_000:.2f}") + "B"
    elif a >= 100_000_000:
        s = f"{round(a / 1_000_000):.0f}M"
    elif a >= 1_000_000:
        s = _strip_zeros(f"{a / 1_000_000:.1f}") + "M"
    elif a >= 1_000:
        s = f"{round(a / 1_000):.0f}K"
    else:
        s = f"{round(a):.0f}"
    return "$" + s


def fmt_num(n) -> str:
    """Comma-formatted integer."""
    try:
        return f"{int(n):,}"
    except (TypeError, ValueError):
        return str(n)


def fmt_date_short(iso: str) -> str:
    """ISO date → 'Apr 27' style."""
    if not iso:
        return "—"
    try:
        d = datetime.strptime(str(iso)[:10], "%Y-%m-%d")
        return f"{d.strftime('%b')} {d.day}"
    except (ValueError, TypeError):
        return str(iso)


def fmt_date_range(start: str, end: str) -> str:
    """'2026-04-19' + '2026-05-03' → 'Apr 19 – May 3, 2026'."""
    try:
        s = datetime.strptime(str(start)[:10], "%Y-%m-%d")
        e = datetime.strptime(str(end)[:10], "%Y-%m-%d")
        if s.year == e.year:
            return f"{s.strftime('%b')} {s.day} – {e.strftime('%b')} {e.day}, {e.year}"
        return (f"{s.strftime('%b')} {s.day}, {s.year} – "
                f"{e.strftime('%b')} {e.day}, {e.year}")
    except (ValueError, TypeError):
        return f"{start} – {end}"


def normalize_investors(value) -> list[str]:
    """Investors may be string ('A, B, C'), list, or None."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(s).strip() for s in value if s and str(s).strip()]
    if isinstance(value, str):
        return [s.strip() for s in value.split(",") if s.strip()]
    return [str(value).strip()]


def bucket_stage(stage) -> str:
    if not stage:
        return "other"
    return STAGE_BUCKETS.get(str(stage).lower().strip(), "other")


def normalize_acquirer_type(t) -> str:
    if not t:
        return "other"
    s = str(t).lower().strip()
    return s if s in ACQUIRER_TYPE_META else "other"


def format_investor_list(investors: list[str], max_chars: int = 80) -> str:
    """Replicates v1's formatInvestors JS: fit names then '+ N more'."""
    if not investors:
        return '<span class="investor-undisc">Undisclosed</span>'
    if len(investors) == 1:
        return esc(investors[0])
    result = esc(investors[0])
    count = 1
    for inv in investors[1:]:
        candidate = result + ", " + esc(inv)
        if len(candidate) > max_chars:
            break
        result = candidate
        count += 1
    if count < len(investors):
        return result + f'<span class="investor-more"> + {len(investors) - count} more</span>'
    return result


# =============================================================================
# FX
# =============================================================================

def fetch_fx_rates(needed_currencies):
    """
    Fixed approximate rates — no network call. Returns (rates_dict, fx_note).

    Converts any currency in FX_RATES_USD; unlisted currencies stay unconverted
    (left out of the USD totals, still shown per-deal). Rates map USD->native,
    so to convert native->USD: divide.
    """
    needed = {c.upper() for c in needed_currencies if c and str(c).upper() != "USD"}
    if not needed:
        return {"USD": 1.0}, None
    note = (f"Capital totals use fixed approximate FX rates ({FX_RATES_AS_OF}); "
            f"per-deal amounts are shown as reported.")
    return dict(FX_RATES_USD), note


def to_usd(value, currency, rates):
    """Native amount → USD. Returns None if not convertible.

    - If rates is None (USD-only mode): only USD records convert; others → None.
    - If currency is missing/None: assume USD (per §5.2 quirk: API sometimes
      returns null currency for normal-looking USD values).
    """
    if value is None:
        return None
    if not currency:
        return float(value)
    code = str(currency).upper()
    if rates is None:
        return float(value) if code == "USD" else None
    rate = rates.get(code)
    if not rate:
        return None
    return float(value) / float(rate)


# =============================================================================
# Disclosure heuristics — see §5.2 / §5.3
# =============================================================================

def is_funding_disclosed(value) -> bool:
    """Funding: 0 and 1 are CatchAll placeholders for undisclosed."""
    return value is not None and value > 1


def is_ma_disclosed(value, display) -> bool:
    """M&A: same as funding, but guard against AUM-extracted-as-deal-value.
    If value exceeds the bogus threshold and display says undisclosed, trust
    the display field and treat as undisclosed."""
    if value is None or value <= 1:
        return False
    if value > DEAL_VALUE_BOGUS_THRESHOLD:
        if display and "undisclosed" in str(display).lower():
            return False
    return True


# =============================================================================
# Funding processing
# =============================================================================

def process_funding(records, rates, fx_failed: bool) -> dict:
    rounds = []
    for r in records:
        e = (r.get("enrichment") or {})
        co = e.get("company_name") or r.get("record_title") or "Unknown"
        sub = e.get("product_description") or ""
        sector = e.get("industry_vertical")
        stage_norm = e.get("funding_stage_normalized")
        stage_bucket = bucket_stage(stage_norm)

        amt_raw = e.get("funding_amount_value")
        currency = (e.get("funding_amount_currency") or "USD")
        amt_display = e.get("funding_amount_display") or "Undisclosed"
        date = e.get("announcement_date")

        cits = r.get("citations") or []
        url = cits[0].get("link") if cits else None

        investors = normalize_investors(e.get("investors"))

        disclosed = is_funding_disclosed(amt_raw)
        if not disclosed:
            amt_usd = None
            amt_tag = None
            amt_disp = "Undisclosed" if amt_display in (None, "") else amt_display
        else:
            amt_usd = to_usd(amt_raw, currency, rates)
            code = str(currency or "USD").upper()
            amt_disp = amt_display
            if code != "USD":
                if amt_usd is not None:
                    amt_tag = f"≈ {fmt_usd(amt_usd)}"
                else:
                    amt_tag = "FX unavailable"
            else:
                amt_tag = None

        # Stage display label: prefer original funding_round, fall back to normalized
        stage_label = (e.get("funding_round")
                       or (str(stage_norm).replace("_", " ").title() if stage_norm else "—"))

        rounds.append({
            "co": co,
            "sub": sub,
            "sector": sector or "",
            "stage_bucket": stage_bucket,
            "stage_label": stage_label,
            "amt": amt_usd,            # USD-equivalent number, for sort/totals
            "amt_display": amt_disp,   # native string ($60M, €2.5M, "Over $4M", "Undisclosed")
            "amt_tag": amt_tag,        # "≈ $X.XM" for non-USD; None otherwise
            "investors": investors,
            "date": date or "",
            "url": url or "",
            "currency": str(currency or "USD").upper(),
            "disclosed": disclosed,
            "excluded": fx_failed and disclosed and str(currency or "USD").upper() != "USD",
        })

    total_deals = len(rounds)
    disclosed_rounds = [r for r in rounds if r["amt"] is not None]
    disclosed_count = sum(1 for r in rounds if r["disclosed"])
    total_amount_usd = sum(r["amt"] for r in disclosed_rounds)

    mega_rounds = [r for r in disclosed_rounds if r["amt"] >= MEGA_ROUND_THRESHOLD]
    mega_count = len(mega_rounds)
    mega_pct = (sum(r["amt"] for r in mega_rounds) / total_amount_usd * 100
                if total_amount_usd > 0 else 0.0)

    deal_stage = {"early": 0, "mid": 0, "late": 0, "other": 0}
    for r in rounds:
        deal_stage[r["stage_bucket"]] += 1

    inv_counts: dict[str, int] = {}
    for r in rounds:
        for inv in r["investors"]:
            inv_counts[inv] = inv_counts.get(inv, 0) + 1
    sorted_inv = sorted(inv_counts.items(), key=lambda x: (-x[1], x[0].lower()))
    most_active = [{"name": k, "deals": v} for k, v in sorted_inv[:3]]
    other_inv_count = max(0, len(inv_counts) - 3)

    sec_counts: dict[str, int] = {}
    sec_totals: dict[str, float] = {}
    sec_canonical: dict[str, str] = {}
    for r in rounds:
        sec = (r["sector"] or "").strip()
        if not sec:
            continue
        key = sec.lower()
        sec_counts[key] = sec_counts.get(key, 0) + 1
        sec_canonical.setdefault(key, sec)
        if r["amt"] is not None:
            sec_totals[key] = sec_totals.get(key, 0.0) + r["amt"]
    sub_sectors = []
    for name, count in sorted(sec_counts.items(), key=lambda x: (-x[1], x[0]))[:6]:
        total = sec_totals.get(name, 0.0)
        sub_sectors.append({
            "name": sec_canonical.get(name, name),
            "count": count,
            "total": total,
            "total_display": fmt_usd(total) if total > 0 else "—",
        })

    return {
        "total_deals": total_deals,
        "total_amount_usd": total_amount_usd,
        "total_amount_display": fmt_usd(total_amount_usd) if total_amount_usd > 0 else "—",
        "disclosed_count": disclosed_count,
        "mega_round_count": mega_count,
        "mega_round_percent": round(mega_pct, 1),
        "deal_stage": deal_stage,
        "most_active_investors": most_active,
        "other_investors_count": other_inv_count,
        "sub_sectors": sub_sectors,
        "rounds": rounds,
    }


# =============================================================================
# M&A processing
# =============================================================================

def process_ma(records, rates, fx_failed: bool) -> dict:
    deals = []
    for r in records:
        e = (r.get("enrichment") or {})
        target = e.get("target_name") or r.get("record_title") or "Unknown"
        target_industry = e.get("target_industry") or ""
        deal_type = e.get("deal_type_display") or "Acquisition"
        acquirer = e.get("acquirer_name") or "Unknown"
        acquirer_type = normalize_acquirer_type(e.get("acquirer_type"))

        amt_raw = e.get("deal_value_value")
        currency = (e.get("deal_value_currency") or "USD")
        amt_display = e.get("deal_value_display") or "Undisclosed"
        date = e.get("announcement_date")

        cits = r.get("citations") or []
        url = cits[0].get("link") if cits else None

        disclosed = is_ma_disclosed(amt_raw, amt_display)
        if not disclosed:
            amt_usd = None
            amt_tag = None
            amt_disp = "Undisclosed" if amt_display in (None, "") else amt_display
        else:
            amt_usd = to_usd(amt_raw, currency, rates)
            code = str(currency or "USD").upper()
            amt_disp = amt_display
            if code != "USD":
                if amt_usd is not None:
                    amt_tag = f"≈ {fmt_usd(amt_usd)}"
                else:
                    amt_tag = "FX unavailable"
            else:
                amt_tag = None

        deals.append({
            "target": target,
            "sub": target_industry,
            "deal_type": deal_type,
            "amt": amt_usd,
            "amt_display": amt_disp,
            "amt_tag": amt_tag,
            "acquirer": acquirer,
            "buyer": acquirer_type,
            "date": date or "",
            "url": url or "",
            "currency": str(currency or "USD").upper(),
            "disclosed": disclosed,
            "excluded": fx_failed and disclosed and str(currency or "USD").upper() != "USD",
        })

    total_deals = len(deals)
    disclosed_deals = [d for d in deals if d["amt"] is not None]
    disclosed_count = sum(1 for d in deals if d["disclosed"])
    total_amount_usd = sum(d["amt"] for d in disclosed_deals)

    largest_deals = sorted(disclosed_deals, key=lambda d: d["amt"], reverse=True)[:3]
    largest = [
        {"acquirer": d["acquirer"], "target": d["target"], "amt_display": d["amt_display"]}
        for d in largest_deals
    ]
    other_deals_count = max(0, total_deals - len(largest))

    acq_counts: dict[str, int] = {}
    for d in deals:
        acq_counts[d["buyer"]] = acq_counts.get(d["buyer"], 0) + 1
    acquirer_type = []
    for key in ["big_tech", "strategic", "financial", "other"]:
        if acq_counts.get(key, 0) > 0:
            acquirer_type.append({"key": key, "count": acq_counts[key]})

    return {
        "total_deals": total_deals,
        "total_amount_usd": total_amount_usd,
        "total_amount_display": fmt_usd(total_amount_usd) if total_amount_usd > 0 else "—",
        "disclosed_count": disclosed_count,
        "largest_deals": largest,
        "other_deals_count": other_deals_count,
        "acquirer_type": acquirer_type,
        "deals": deals,
    }


# =============================================================================
# Capital ratio
# =============================================================================

def compute_capital_ratio(funding: dict, ma: dict) -> dict:
    eq = funding["total_amount_usd"]
    ma_amt = ma["total_amount_usd"]

    if eq <= 0 and ma_amt <= 0:
        return {
            "ratio": 0,
            "headline": "No disclosed capital",
            "subline": "Neither feed disclosed deal values in this window.",
            "ma_bar_pct": 0,
            "equity_bar_pct": 0,
            "ma_amount_display": "—",
            "equity_amount_display": "—",
            "footnote": (f"M&A coverage: {ma['disclosed_count']} of {ma['total_deals']} "
                         f"deals disclose value. Funding coverage: "
                         f"{funding['disclosed_count']} of {funding['total_deals']}."),
        }

    if eq <= 0:
        headline = "M&A only"
        subline = "Disclosed M&A capital with no disclosed equity for comparison."
    elif ma_amt <= 0:
        headline = "Equity only"
        subline = "Disclosed equity capital with no disclosed M&A for comparison."
    else:
        ratio = ma_amt / eq
        if 0.9 <= ratio <= 1.1:
            headline = "M&A ≈ equity"
            subline = "Disclosed M&A capital is roughly balanced against disclosed equity."
        elif ratio > 1:
            headline = f"M&A {ratio:.1f}× equity"
            subline = "Disclosed M&A capital outpaces disclosed equity over the period."
        else:
            inv = 1 / ratio if ratio > 0 else 0
            headline = f"Equity {inv:.1f}× M&A"
            subline = "Disclosed equity capital outpaces disclosed M&A over the period."

    dominant = max(eq, ma_amt)
    if dominant > 0:
        ma_pct = (ma_amt / dominant) * 100
        eq_pct = (eq / dominant) * 100
    else:
        ma_pct = eq_pct = 0

    return {
        "ratio": (ma_amt / eq) if eq > 0 else 0,
        "headline": headline,
        "subline": subline,
        "ma_bar_pct": round(ma_pct, 1),
        "equity_bar_pct": round(eq_pct, 1),
        "ma_amount_display": fmt_usd(ma_amt) if ma_amt > 0 else "—",
        "equity_amount_display": fmt_usd(eq) if eq > 0 else "—",
        "footnote": (f"M&A coverage: {ma['disclosed_count']} of {ma['total_deals']} "
                     f"deals disclose value. Funding coverage: "
                     f"{funding['disclosed_count']} of {funding['total_deals']}."),
    }


# =============================================================================
# HTML rendering — pre-render every chunk Python-side
# =============================================================================

def _stack_segment(width_pct: float, color: str, label: str) -> str:
    label_html = label if width_pct >= 7 else ""
    return (f'<div class="stack-seg" '
            f'style="width:{width_pct:.1f}%;background:{color};">'
            f'{label_html}</div>')


def render_funding_stage_bar(funding: dict) -> str:
    stages = funding["deal_stage"]
    total = sum(stages.values())
    if total == 0:
        return '<div class="stack-seg" style="width:100%;background:#B5B2AA;">No data</div>'
    parts = []
    for key in ("early", "mid", "late", "other"):
        c = stages.get(key, 0)
        pct = (c / total * 100) if total else 0
        parts.append(_stack_segment(pct, DEAL_STAGE_META[key]["color"], f"{round(pct)}%"))
    return "".join(parts)


def render_funding_stage_legend(funding: dict) -> str:
    stages = funding["deal_stage"]
    rows = []
    for key in ("early", "mid", "late", "other"):
        meta = DEAL_STAGE_META[key]
        c = stages.get(key, 0)
        rows.append(
            f'<div class="legend-item">'
            f'<span><span class="legend-dot" style="background:{meta["color"]};"></span>'
            f'{esc(meta["label"])} '
            f'<span style="color:var(--color-text-tertiary, #8d8a78);font-size:11px;">'
            f'{esc(meta["sublabel"])}</span></span>'
            f'<span style="color:var(--color-text-secondary, #5d5b51);">'
            f'{c} deal{"s" if c != 1 else ""}</span></div>'
        )
    return "".join(rows)


def render_funding_investors(funding: dict) -> str:
    rows = []
    for inv in funding["most_active_investors"]:
        rows.append(
            f'<div class="topdeal-row">'
            f'<span class="topdeal-name">{esc(inv["name"])}</span>'
            f'<span class="topdeal-amt">{inv["deals"]} '
            f'deal{"s" if inv["deals"] != 1 else ""}</span></div>'
        )
    if not rows:
        rows.append('<div class="topdeal-row"><span class="topdeal-name" '
                    'style="color:var(--color-text-tertiary, #8d8a78);font-style:italic;">'
                    'No named investors disclosed.</span></div>')
    return "".join(rows)


def render_funding_subsectors(funding: dict) -> str:
    secs = funding["sub_sectors"]
    if not secs:
        return ('<div class="cat-row"><span class="cat-label" '
                'style="color:var(--color-text-tertiary, #8d8a78);font-style:italic;">'
                'No sub-sectors detected.</span><span></span><span></span></div>')
    max_count = max((s["count"] for s in secs), default=1)
    rows = []
    for s in secs:
        bar_pct = (s["count"] / max_count * 100) if max_count else 0
        rows.append(
            f'<div class="cat-row">'
            f'<span class="cat-label">{esc(s["name"])}</span>'
            f'<div class="cat-bar-track"><div class="cat-bar-fill" '
            f'style="width:{bar_pct:.1f}%;">{s["count"]}</div></div>'
            f'<span class="cat-amt">{esc(s["total_display"])}</span></div>'
        )
    return "".join(rows)


def render_ma_acquirer_bar(ma: dict) -> str:
    types = ma["acquirer_type"]
    total = sum(t["count"] for t in types)
    if total == 0:
        return '<div class="stack-seg" style="width:100%;background:#B5B2AA;">No data</div>'
    parts = []
    for t in types:
        meta = ACQUIRER_TYPE_META.get(t["key"], ACQUIRER_TYPE_META["other"])
        pct = (t["count"] / total * 100) if total else 0
        parts.append(_stack_segment(pct, meta["color"], f"{round(pct)}%"))
    return "".join(parts)


def render_ma_acquirer_legend(ma: dict) -> str:
    rows = []
    for t in ma["acquirer_type"]:
        meta = ACQUIRER_TYPE_META.get(t["key"], ACQUIRER_TYPE_META["other"])
        rows.append(
            f'<div class="legend-item">'
            f'<span><span class="legend-dot" style="background:{meta["color"]};"></span>'
            f'{esc(meta["label"])}</span>'
            f'<span style="color:var(--color-text-secondary, #5d5b51);">'
            f'{t["count"]} deal{"s" if t["count"] != 1 else ""}</span></div>'
        )
    if not rows:
        rows.append('<div class="legend-item"><span style="color:var(--color-text-tertiary, #8d8a78);'
                    'font-style:italic;">No acquirer-type data.</span></div>')
    return "".join(rows)


def render_ma_largest_deals(ma: dict) -> str:
    rows = []
    for d in ma["largest_deals"]:
        rows.append(
            f'<div class="topdeal-row">'
            f'<span class="topdeal-name">{esc(d["acquirer"])} '
            f'<span class="arrow">→</span> {esc(d["target"])}</span>'
            f'<span class="topdeal-amt">{esc(d["amt_display"])}</span></div>'
        )
    if not rows:
        rows.append('<div class="topdeal-row"><span class="topdeal-name" '
                    'style="color:var(--color-text-tertiary, #8d8a78);font-style:italic;">'
                    'No deal values disclosed.</span></div>')
    return "".join(rows)


def render_funding_rows(funding: dict) -> str:
    rows = []
    for r in funding["rounds"]:
        # data-amt for sort (-1 sentinel for undisclosed so they sort last)
        # Filter/sort by company-name use first cell's textContent — saves ~200ch/row
        data_amt = f"{r['amt']:.0f}" if r["amt"] is not None else "-1"
        data_date = r["date"] or ""

        amt_html = (
            f'<span class="amt amt-undisc">{esc(r["amt_display"] or "Undisclosed")}</span>'
            if r["amt"] is None
            else (f'<span class="amt">{esc(r["amt_display"])}</span>' +
                  (f'<span class="amt-tag">{esc(r["amt_tag"])}</span>' if r["amt_tag"] else ""))
        )

        sector_chip = (f'<div><span class="sector-chip">{esc(r["sector"])}</span></div>'
                       if r["sector"] else "")
        full_inv = ", ".join(r["investors"])
        inv_html = format_investor_list(r["investors"], 80)
        date_str = fmt_date_short(r["date"])
        url_html = (f'<a href="{esc(r["url"])}" class="src-link" target="_blank" '
                    f'rel="noopener">↗</a>') if r["url"] else ""

        # title attrs only when text might overflow CSS clamp / ellipsis
        co_title = f' title="{esc(r["co"])}"' if r["co"] and len(r["co"]) > 18 else ""
        sub_title = f' title="{esc(r["sub"])}"' if r["sub"] and len(r["sub"]) > 60 else ""
        inv_title = f' title="{esc(full_inv)}"' if full_inv and len(full_inv) > 40 else ""

        rows.append(
            f'<tr data-amt="{data_amt}" data-date="{esc(data_date)}">'
            f'<td><div class="co-name"{co_title}>{esc(r["co"])}</div>'
            f'<div class="co-sub"{sub_title}>{esc(r["sub"])}</div>'
            f'{sector_chip}</td>'
            f'<td class="right">{amt_html}</td>'
            f'<td><span class="stage-cell">{esc(r["stage_label"])}</span></td>'
            f'<td><div class="investors-list"{inv_title}>{inv_html}</div></td>'
            f'<td><span class="date-cell">{date_str}</span></td>'
            f'<td>{url_html}</td>'
            f'</tr>'
        )
    return "".join(rows)


def render_ma_rows(ma: dict) -> str:
    rows = []
    for d in ma["deals"]:
        data_amt = f"{d['amt']:.0f}" if d["amt"] is not None else "-1"
        data_date = d["date"] or ""

        amt_html = (
            f'<span class="amt amt-undisc">{esc(d["amt_display"] or "Undisclosed")}</span>'
            if d["amt"] is None
            else (f'<span class="amt">{esc(d["amt_display"])}</span>' +
                  (f'<span class="amt-tag">{esc(d["amt_tag"])}</span>' if d["amt_tag"] else ""))
        )
        meta = ACQUIRER_TYPE_META.get(d["buyer"], ACQUIRER_TYPE_META["other"])
        date_str = fmt_date_short(d["date"])
        url_html = (f'<a href="{esc(d["url"])}" class="src-link" target="_blank" '
                    f'rel="noopener">↗</a>') if d["url"] else ""

        target_title = f' title="{esc(d["target"])}"' if d["target"] and len(d["target"]) > 18 else ""
        sub_title = f' title="{esc(d["sub"])}"' if d["sub"] and len(d["sub"]) > 30 else ""
        acq_title = f' title="{esc(d["acquirer"])}"' if d["acquirer"] and len(d["acquirer"]) > 25 else ""

        rows.append(
            f'<tr data-amt="{data_amt}" data-date="{esc(data_date)}">'
            f'<td><div class="co-name"{target_title}>{esc(d["target"])}</div>'
            f'<div class="co-sub"{sub_title}>{esc(d["sub"])}</div></td>'
            f'<td><span class="stage-cell">{esc(d["deal_type"])}</span></td>'
            f'<td class="right">{amt_html}</td>'
            f'<td><div class="lead-name-strong"{acq_title}>{esc(d["acquirer"])}</div>'
            f'<span class="buyer-chip {meta["chipClass"]}">{esc(meta["label"])}</span></td>'
            f'<td><span class="date-cell">{date_str}</span></td>'
            f'<td>{url_html}</td>'
            f'</tr>'
        )
    return "".join(rows)


def render_html(input_data: dict, template_text: str) -> str:
    meta = input_data.get("meta") or {}
    funding_pull = input_data.get("funding_pull") or {}
    ma_pull = input_data.get("ma_pull") or {}

    funding_records = funding_pull.get("all_records") or []
    ma_records = ma_pull.get("all_records") or []

    # FX
    needed_currencies = set()
    for r in funding_records:
        e = r.get("enrichment") or {}
        if e.get("funding_amount_currency"):
            needed_currencies.add(e["funding_amount_currency"])
    for r in ma_records:
        e = r.get("enrichment") or {}
        if e.get("deal_value_currency"):
            needed_currencies.add(e["deal_value_currency"])

    rates, fx_note = fetch_fx_rates(needed_currencies)
    fx_failed = (rates is None)
    if fx_failed:
        excluded = sum(
            1 for r in funding_records
            if (r.get("enrichment") or {}).get("funding_amount_currency", "USD")
            and str((r.get("enrichment") or {}).get("funding_amount_currency", "USD")).upper() != "USD"
            and is_funding_disclosed((r.get("enrichment") or {}).get("funding_amount_value"))
        ) + sum(
            1 for r in ma_records
            if (r.get("enrichment") or {}).get("deal_value_currency", "USD")
            and str((r.get("enrichment") or {}).get("deal_value_currency", "USD")).upper() != "USD"
            and is_ma_disclosed(
                (r.get("enrichment") or {}).get("deal_value_value"),
                (r.get("enrichment") or {}).get("deal_value_display"),
            )
        )
        fx_note = (f"FX unavailable — {excluded} non-USD deal{'s' if excluded != 1 else ''} "
                   f"excluded from totals. Reply 'retry' to retry conversion.")

    funding = process_funding(funding_records, rates, fx_failed)
    ma = process_ma(ma_records, rates, fx_failed)
    ratio = compute_capital_ratio(funding, ma)

    # Run stats
    articles_scanned = (funding_pull.get("candidate_records") or 0) + (ma_pull.get("candidate_records") or 0)
    deals_found = (funding_pull.get("valid_records") or funding["total_deals"]) + \
                  (ma_pull.get("valid_records") or ma["total_deals"])

    fx_block = (f'<div style="font-size:11px;color:var(--color-text-tertiary, #8d8a78);'
                f'font-style:italic;margin-top:1.25rem;line-height:1.45;">'
                f'{esc(fx_note)}</div>') if fx_note else ""

    ratio_footnote = (f'<div class="footnote">{esc(ratio["footnote"])}</div>'
                      if ratio.get("footnote") else "")

    replacements = {
        "MARKET":            esc(meta.get("market", "")),
        "LOCATION":          esc(meta.get("location", "")),
        "WINDOW_DAYS":       esc(meta.get("window_days", "")),
        "DATE_RANGE":        esc(fmt_date_range(meta.get("start_date", ""), meta.get("end_date", ""))),
        "ARTICLES_SCANNED":  fmt_num(articles_scanned),
        "DEALS_FOUND":       fmt_num(deals_found),

        "FUNDING_TOTAL_DEALS":     str(funding["total_deals"]),
        "FUNDING_TOTAL_AMOUNT":    esc(funding["total_amount_display"]),
        "FUNDING_DISCLOSED_COUNT": str(funding["disclosed_count"]),
        "FUNDING_STAGE_BAR":       render_funding_stage_bar(funding),
        "FUNDING_STAGE_LEGEND":    render_funding_stage_legend(funding),
        "FUNDING_INVESTORS":       render_funding_investors(funding),
        "FUNDING_OTHER_INVESTORS": str(funding["other_investors_count"]),
        "FUNDING_INVESTOR_PLURAL": "s" if funding["other_investors_count"] != 1 else "",
        "FUNDING_SUBSECTORS":      render_funding_subsectors(funding),

        "MA_TOTAL_DEALS":     str(ma["total_deals"]),
        "MA_TOTAL_AMOUNT":    esc(ma["total_amount_display"]),
        "MA_DISCLOSED_COUNT": str(ma["disclosed_count"]),
        "MA_ACQUIRER_BAR":    render_ma_acquirer_bar(ma),
        "MA_ACQUIRER_LEGEND": render_ma_acquirer_legend(ma),
        "MA_LARGEST_DEALS":   render_ma_largest_deals(ma),
        "MA_OTHER_DEALS":     str(ma["other_deals_count"]),
        "MA_DEAL_PLURAL":     "s" if ma["other_deals_count"] != 1 else "",

        "RATIO_HEADLINE":         esc(ratio["headline"]),
        "RATIO_SUBLINE":          esc(ratio["subline"]),
        "RATIO_MA_PCT":           f"{ratio['ma_bar_pct']:.1f}",
        "RATIO_EQUITY_PCT":       f"{ratio['equity_bar_pct']:.1f}",
        "RATIO_MA_AMOUNT":        esc(ratio["ma_amount_display"]),
        "RATIO_EQUITY_AMOUNT":    esc(ratio["equity_amount_display"]),
        "RATIO_FOOTNOTE":         ratio_footnote,

        "FUNDING_TAB_COUNT": str(funding["total_deals"]),
        "MA_TAB_COUNT":      str(ma["total_deals"]),
        "FUNDING_ROWS":      render_funding_rows(funding),
        "MA_ROWS":           render_ma_rows(ma),

        "FX_NOTE_BLOCK":     fx_block,
    }

    if input_data.get("ma_pending"):
        replacements["MA_TOTAL_DEALS"]     = "—"
        replacements["MA_TAB_COUNT"]       = "—"
        replacements["MA_TOTAL_AMOUNT"]    = "Searching…"
        replacements["MA_DISCLOSED_COUNT"] = "—"
        replacements["RATIO_HEADLINE"]     = "M&A still searching"
        replacements["RATIO_SUBLINE"]      = (
            "Funding is shown below. The M&A search is still running — "
            "reply “refresh” once it finishes to complete the capital view."
        )
        replacements["RATIO_FOOTNOTE"]     = ""

    return substitute(template_text, replacements)


def substitute(template: str, replacements: dict) -> str:
    out = template
    for key, value in replacements.items():
        out = out.replace("{{" + key + "}}", str(value))
    return out


# =============================================================================
# Minify
# =============================================================================

def minify_html(s: str) -> str:
    """Regex minify: strip HTML/CSS comments, collapse whitespace, and
    compress CSS rule whitespace (around delimiters).

    Preserves URLs and content-meaningful single spaces. Skips JS line
    comments — we don't write any, and URLs contain //."""
    s = re.sub(r"<!--(?!\[if).*?-->", "", s, flags=re.DOTALL)
    s = re.sub(r"/\*.*?\*/", "", s, flags=re.DOTALL)

    # Compress CSS inside <style> blocks: tighten whitespace around
    # delimiters. Preserves selectors, class names, and rule structure —
    # no aggressive class-name shortening per §3.8.
    def _compress_css(m):
        css = m.group(1)
        css = re.sub(r"\s+", " ", css)
        css = re.sub(r"\s*([{};:,])\s*", r"\1", css)
        css = re.sub(r";}", "}", css)
        return f"<style>{css}</style>"
    s = re.sub(r"<style>(.*?)</style>", _compress_css, s, flags=re.DOTALL)

    # Collapse remaining whitespace runs and trim space between tags.
    s = re.sub(r"[ \t\r\n]+", " ", s)
    s = re.sub(r">\s+<", "><", s)
    return s.strip()


# =============================================================================
# Markdown rendering — Tier 3 fallback
# =============================================================================

def render_markdown(input_data: dict) -> str:
    meta = input_data.get("meta") or {}
    funding_pull = input_data.get("funding_pull") or {}
    ma_pull = input_data.get("ma_pull") or {}
    funding_records = funding_pull.get("all_records") or []
    ma_records = ma_pull.get("all_records") or []

    needed_currencies = set()
    for r in funding_records:
        e = r.get("enrichment") or {}
        if e.get("funding_amount_currency"):
            needed_currencies.add(e["funding_amount_currency"])
    for r in ma_records:
        e = r.get("enrichment") or {}
        if e.get("deal_value_currency"):
            needed_currencies.add(e["deal_value_currency"])
    rates, fx_note = fetch_fx_rates(needed_currencies)
    fx_failed = (rates is None)

    funding = process_funding(funding_records, rates, fx_failed)
    ma = process_ma(ma_records, rates, fx_failed)
    ratio = compute_capital_ratio(funding, ma)

    out = []
    out.append(f"# VC pack — {meta.get('market', '')}, {meta.get('location', '')}")
    out.append(f"_{fmt_date_range(meta.get('start_date', ''), meta.get('end_date', ''))}_  "
               f"({meta.get('window_days', '?')}-day window)")
    out.append("")
    out.append(f"**{funding['total_amount_display']}** raised across "
               f"**{funding['total_deals']}** funding rounds "
               f"({funding['disclosed_count']} disclose value).  ")
    if input_data.get("ma_pending"):
        out.append("_M&A search still running — reply 'refresh' once it "
                   "finishes to complete the dashboard._  ")
        out.append("_M&A still searching._ Funding is shown below; the capital "
                   "view completes once the M&A search finishes.")
    else:
        out.append(f"**{ma['total_amount_display']}** disclosed M&A across "
                   f"**{ma['total_deals']}** deals "
                   f"({ma['disclosed_count']} disclose value).  ")
        out.append(f"_{ratio['headline']}._ {ratio['subline']}")
    out.append("")
    out.append("## Funding rounds")
    out.append("")
    out.append("| Company | Amount | Stage | Investors | Date |")
    out.append("|---|---:|---|---|---|")
    rows_sorted = sorted(funding["rounds"],
                         key=lambda r: (r["amt"] if r["amt"] is not None else -1),
                         reverse=True)
    for r in rows_sorted:
        amt = r["amt_display"]
        if r["amt_tag"]:
            amt = f"{amt} ({r['amt_tag']})"
        invs = ", ".join(r["investors"][:3])
        if len(r["investors"]) > 3:
            invs += f" + {len(r['investors']) - 3} more"
        out.append(
            f"| {_md_cell(r['co'])} | {_md_cell(amt)} | "
            f"{_md_cell(r['stage_label'])} | {_md_cell(invs or 'Undisclosed')} | "
            f"{_md_cell(fmt_date_short(r['date']))} |"
        )
    out.append("")
    out.append("## Acquisitions")
    out.append("")
    if input_data.get("ma_pending"):
        out.append("_Still searching — reply 'refresh' to load M&A results._")
    else:
        out.append("| Target | Deal type | Value | Acquirer | Date |")
        out.append("|---|---|---:|---|---|")
        deals_sorted = sorted(ma["deals"],
                              key=lambda d: (d["amt"] if d["amt"] is not None else -1),
                              reverse=True)
        for d in deals_sorted:
            amt = d["amt_display"]
            if d["amt_tag"]:
                amt = f"{amt} ({d['amt_tag']})"
            buyer = ACQUIRER_TYPE_META.get(d["buyer"], ACQUIRER_TYPE_META["other"])["label"]
            out.append(
                f"| {_md_cell(d['target'])} | {_md_cell(d['deal_type'])} | "
                f"{_md_cell(amt)} | {_md_cell(d['acquirer'])} ({buyer}) | "
                f"{_md_cell(fmt_date_short(d['date']))} |"
            )
    out.append("")
    if fx_note:
        out.append(f"_{fx_note}_")
    if fx_failed:
        out.append(f"_FX unavailable — non-USD deals excluded from totals._")

    articles = (funding_pull.get("candidate_records") or 0) + (ma_pull.get("candidate_records") or 0)
    deals_found = (funding_pull.get("valid_records") or funding["total_deals"]) + \
                  (ma_pull.get("valid_records") or ma["total_deals"])
    out.append("")
    out.append(f"_{fmt_num(articles)} articles scanned → {fmt_num(deals_found)} deals "
               f"found · via CatchAll_")
    return "\n".join(out)


def _md_cell(s) -> str:
    """Escape cell content for markdown tables: pipes and newlines."""
    if s is None:
        return ""
    return str(s).replace("|", "\\|").replace("\n", " ").strip()


# =============================================================================
# Input acquisition — saved pull files or a combined input.json (no network)
# =============================================================================

def load_pull_file(path: str) -> dict:
    """Load a CatchAll pull from disk. Auto-unwraps MCP-saved tool
    results of the form {"result": "<inner JSON string>"}.

    Use this when the agent's environment saves large MCP responses to
    disk — render.py reads them directly instead of having Claude re-emit
    the contents through token generation."""
    text = Path(path).read_text(encoding="utf-8")
    data = json.loads(text)
    if isinstance(data, dict) and isinstance(data.get("result"), str):
        try:
            data = json.loads(data["result"])
        except json.JSONDecodeError:
            pass  # fall through with the unwrapped {result: ...}
    return data


def build_input_data(args: argparse.Namespace) -> dict:
    """Assemble the input dict from whichever mode the caller supplied.

    Modes (in priority order):
      1. --funding-pull + --ma-pull : load saved pull files
      2. positional input.json      : pre-assembled combined input
    """
    ma_pending = bool(getattr(args, "ma_pending", False))
    if args.funding_pull and (args.ma_pull or ma_pending):
        funding = load_pull_file(args.funding_pull)
        ma = load_pull_file(args.ma_pull) if args.ma_pull else {}
    elif args.input:
        data = json.loads(Path(args.input).read_text(encoding="utf-8"))
        if ma_pending:
            data["ma_pending"] = True
        return data
    else:
        sys.exit(
            "vc-pack: provide one of:\n"
            "  --funding-pull P1  --ma-pull P2   (saved pull files)\n"
            "  positional input.json             (pre-assembled combined input)"
        )

    dr = funding.get("date_range") or {}
    start_date = (args.start_date
                  or (dr.get("start_date") or "")[:10])
    end_date = (args.end_date
                or (dr.get("end_date") or "")[:10])

    return {
        "funding_pull": {
            "candidate_records": funding.get("candidate_records", 0),
            "valid_records": funding.get("valid_records", 0),
            "all_records": funding.get("all_records") or [],
        },
        "ma_pull": {
            "candidate_records": ma.get("candidate_records", 0),
            "valid_records": ma.get("valid_records", 0),
            "all_records": ma.get("all_records") or [],
        },
        "meta": {
            "market": args.market,
            "location": args.location,
            "window_days": args.window_days,
            "start_date": start_date,
            "end_date": end_date,
        },
        "ma_pending": ma_pending,
    }


# =============================================================================
# Main
# =============================================================================

def _process_feeds(input_data):
    """Process both feeds (FX + parse) for the data-file export. Mirrors what
    the renderers do internally, kept separate so the render path is untouched."""
    funding_records = (input_data.get("funding_pull") or {}).get("all_records") or []
    ma_records = (input_data.get("ma_pull") or {}).get("all_records") or []
    needed = set()
    for r in funding_records:
        c = (r.get("enrichment") or {}).get("funding_amount_currency")
        if c:
            needed.add(c)
    for r in ma_records:
        c = (r.get("enrichment") or {}).get("deal_value_currency")
        if c:
            needed.add(c)
    rates, _ = fetch_fx_rates(needed)
    fx_failed = (rates is None)
    return (process_funding(funding_records, rates, fx_failed),
            process_ma(ma_records, rates, fx_failed))


def write_data_files(meta, funding, ma, out_prefix, ma_pending=False):
    """Write the VC pack's underlying records as <prefix>.json / .csv / .xlsx.

    Returns the list of paths written (the xlsx is skipped, with a stderr note,
    if openpyxl is not installed)."""
    meta = meta or {}
    json_path = out_prefix + ".json"
    csv_path = out_prefix + ".csv"
    xlsx_path = out_prefix + ".xlsx"

    funding_out = [{
        "company": r["co"], "description": r["sub"], "sector": r["sector"],
        "stage": r["stage_label"], "amount_display": r["amt_display"],
        "amount_usd": r["amt"], "currency": r["currency"],
        "investors": r["investors"], "date": r["date"], "url": r["url"],
    } for r in funding["rounds"]]
    ma_out = [{
        "target": d["target"], "target_industry": d["sub"],
        "deal_type": d["deal_type"], "acquirer": d["acquirer"],
        "acquirer_type": d["buyer"], "amount_display": d["amt_display"],
        "amount_usd": d["amt"], "currency": d["currency"],
        "date": d["date"], "url": d["url"],
    } for d in ma["deals"]]

    data = {
        "meta": {
            **meta,
            "skill": "vc-pack",
            "mode": "base",
            "ma_pending": bool(ma_pending),
            "recall": {
                "funding_deals": funding["total_deals"],
                "ma_deals": ma["total_deals"],
            },
        },
        "funding": funding_out,
        "ma": ma_out,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["feed", "company_or_target", "counterparty",
                    "counterparty_type", "amount", "stage_or_deal_type",
                    "sector", "date", "url"])
        for r in funding["rounds"]:
            w.writerow(["funding", r["co"], "; ".join(r["investors"]),
                        "investors", r["amt_display"], r["stage_label"],
                        r["sector"], r["date"], r["url"]])
        for d in ma["deals"]:
            w.writerow(["ma", d["target"], d["acquirer"], d["buyer"],
                        d["amt_display"], d["deal_type"], d["sub"],
                        d["date"], d["url"]])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        sys.stderr.write("vc-pack: openpyxl not installed — wrote JSON + CSV "
                         "only. Run `pip install openpyxl --break-system-packages` "
                         "for the xlsx.\n")
        return [json_path, csv_path]

    wb = Workbook()
    ov = wb.active
    ov.title = "Overview"
    ov["A1"] = f"VC pack — {meta.get('market', '')}, {meta.get('location', '')}"
    ov["A2"] = f"{meta.get('window_days', '?')}-day window · raw dataset behind the dashboard"
    ov["A4"] = "Funding = capital into the market; M&A = capital out. One row per record on each sheet."
    ov["A6"] = "More with CatchAll:"
    links_path = Path(__file__).resolve().parent.parent / "references" / "links.json"
    links = json.loads(links_path.read_text(encoding="utf-8"))
    _row = 7
    for link in links["footer_links"]:
        if link.get("watchlist_only"):
            continue  # vc-pack is not a watchlist skill
        ov[f"A{_row}"] = f"{link['label']} — {link['url']}"
        _row += 1
    ov[f"A{_row}"] = f"Questions? {links['support_email']}"

    navy = PatternFill("solid", fgColor="1F3A5F")
    white = Font(color="FFFFFF", bold=True)

    fs = wb.create_sheet("Funding")
    fs.append(["company", "amount", "stage", "sector", "investors", "date", "url"])
    for r in funding["rounds"]:
        fs.append([r["co"], r["amt_display"], r["stage_label"], r["sector"],
                   "; ".join(r["investors"]), r["date"], r["url"]])

    ms = wb.create_sheet("M&A")
    ms.append(["target", "acquirer", "acquirer_type", "deal_type",
               "amount", "target_industry", "date", "url"])
    for d in ma["deals"]:
        ms.append([d["target"], d["acquirer"], d["buyer"], d["deal_type"],
                   d["amt_display"], d["sub"], d["date"], d["url"]])

    for sheet in (fs, ms):
        for cell in sheet[1]:
            cell.fill = navy
            cell.font = white
        sheet.freeze_panes = "A2"

    wb.save(xlsx_path)
    return [xlsx_path, json_path, csv_path]


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Render a VC pack dashboard.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Two input modes:\n"
            "  1. --funding-pull P1  --ma-pull P2   (saved pulls; auto-unwraps MCP)\n"
            "  2. <input.json>                      (pre-assembled combined input)"
        ),
    )
    parser.add_argument("input", nargs="?", default=None,
                        help="Path to combined input JSON (mode 2).")
    parser.add_argument("--funding-pull",
                        help="Path to saved funding pull (raw or MCP-wrapped).")
    parser.add_argument("--ma-pull",
                        help="Path to saved M&A pull (raw or MCP-wrapped).")
    parser.add_argument("--market", default="Unknown")
    parser.add_argument("--location", default="global")
    parser.add_argument("--window-days", type=int, default=30)
    parser.add_argument("--start-date", default=None)
    parser.add_argument("--end-date", default=None)
    parser.add_argument("--format", choices=["html", "markdown"], default="html")
    parser.add_argument("--output", default=None,
                        help="Write to PATH instead of stdout.")
    parser.add_argument("--no-minify", action="store_true")
    parser.add_argument("--ma-pending", action="store_true",
                        help="Render funding now with M&A marked still-running "
                             "(use when the M&A feed hasn't completed by the cap).")
    parser.add_argument("--data-out", default=None,
                        help="Also write the funding+M&A records as PREFIX.json / "
                             ".csv / .xlsx (the downloadable dataset behind the dashboard).")
    args = parser.parse_args(argv)

    input_data = build_input_data(args)

    if args.data_out:
        funding, ma = _process_feeds(input_data)
        paths = write_data_files(input_data.get("meta") or {}, funding, ma,
                                 args.data_out, bool(input_data.get("ma_pending")))
        sys.stderr.write("vc-pack data files: " + ", ".join(paths) + "\n")

    if args.format == "markdown":
        output = render_markdown(input_data)
    else:
        template_path = Path(__file__).resolve().parent.parent / "assets" / "dashboard.html"
        template_text = template_path.read_text(encoding="utf-8")
        output = render_html(input_data, template_text)
        if not args.no_minify:
            output = minify_html(output)

    if args.output:
        Path(args.output).write_text(output, encoding="utf-8")
    else:
        sys.stdout.write(output)
        if not output.endswith("\n"):
            sys.stdout.write("\n")


if __name__ == "__main__":
    main()
